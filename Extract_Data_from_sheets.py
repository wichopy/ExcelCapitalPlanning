import os
import openpyxl as opxl
import click
import Tkinter
import tkFileDialog
import csv
import logging
import traceback
import shutil
"""
'Go through each excel file in directory you choose.
'Extract values from a certain sheet and range.
'Store these files and load to a csv file once all spreadsheets are checked.
"""
root = Tkinter.Tk()
root.withdraw()

rootfolder = tkFileDialog.askdirectory(title='Open file to check')
print ("Will analyze files at directory: "+rootfolder)
Building_lvl_info = []
count=0
countroot = 0
for dirName, subdirlist, fileList in os.walk(rootfolder):
    if countroot > 0:
        break
    print("Checking directory: "+dirName)
    for filename in fileList:
        print filename
        try:
            if filename.endswith(".xlsx"):
                workbook = opxl.load_workbook(dirName+"\\"+filename)
                if "checks" in workbook.get_sheet_names():
                    report = workbook.get_sheet_by_name(u'checks')
                    BuildingID = filename[:7]
                    BuildingType = report.cell(row = 16, column = 3).value
                    ConstructionClass = report.cell(row = 16, column = 4).value
                    Exteriorcondition = report.cell(row = 16, column = 5).value
                    
                    Building_lvl_info.append([BuildingID,BuildingType,ConstructionClass, Exteriorcondition, filename])
                    print Building_lvl_info[count]
                    count=count+1
        except Exception as e:
            Building_lvl_info.append(["error",0,0,0,filename])
            continue
    countroot += 1                
            
with open("BuildingLevelData.csv","wb") as f:
    writer = csv.writer(f)
    writer.writerows(Building_lvl_info)
