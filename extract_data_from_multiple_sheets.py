"""
This script was written as a request from my manager to extract snippets of data from multiple sheets within a spreadsheet.

The process would have taken a unnecessary long amount of time to do manually.

No process was originally put in place to have data from the excel sheet filled in by field personnel to be extracted manually and 
I had to come up with a quick automated way to do it. 

Luckily the data that needed to be extracted were in the same cells for all 500+ sheets.

I used the openpyxl module to extract data from the sheets with cell references, and combined all the data into a dataframe in pandas
for export to excel.

"""
import pandas as pd
import os
import csv
import os.path,time
import datetime
import openpyxl as opxl
import openpyxl.utils as utils

print "Start extracting from reports."
datenow = str(datetime.datetime.now().strftime('%Y%m%d'))

fname_list = []
#sheetname = BCA Environmental Checklist
enviro_list = []

#sheetname = "checks"
bldg_list = []

questdir = 'C:\\Users\\chouw\\Desktop\\reviewed\\'

#functions for cell refernces.
def bldg_data_pull(wb):

    if "checks" in wb.get_sheet_names():
        report = wb.get_sheet_by_name(u'checks')
        BuildingType = report.cell(row = 17, column = 3).value
        ConstructionClass = report.cell(row = 17, column = 4).value
        Exteriorcondition = report.cell(row = 17, column = 5).value
        return [BuildingType,ConstructionClass, Exteriorcondition]
    else:
        return ["did not have checks tab"]*3
    
def enviro_data_pull(wb):

    colA = utils.column_index_from_string('A')
    colH = utils.column_index_from_string('H')
    if "BCA Environmental Checklist" in wb.get_sheet_names():
        report = wb.get_sheet_by_name(u'BCA Environmental Checklist')
        if report.cell(row = 15, column = colH).value is None:
			#some field guys decided to highlight the cell red/green rathre then fill in a Y or N..
            q1 = report.cell(row = 15, column = colH).fill
        else:
            q1 = report.cell(row = 15, column = colH).value
            
        q1_com = report.cell(row = 16, column = colA).value
        
        if report.cell(row = 17, column = colH).value is None:
            q2 = report.cell(row = 17, column = colH).fill
        else:
            q2 = report.cell(row = 17, column = colH).value
            
        q2_com = report.cell(row = 18, column = colA).value
        
        if report.cell(row = 19, column = colH).value is None:
            q3 = report.cell(row = 19, column = colH).fill
        else:
            q3 = report.cell(row = 19, column = colH).value     
        
        q3_com = report.cell(row = 20, column = colA).value
        
        if report.cell(row = 21, column = colH).value is None:
            q4 = report.cell(row = 21, column = colH).fill
        else:
            q4 = report.cell(row = 21, column = colH).value          
        
        q4_com = report.cell(row = 22, column = colA).value
        
        add_comm = report.cell(row = 23, column = colA).value
        return [q1,q1_com,q2,q2_com,q3,q3_com,q4,q4_com,add_comm]
    else:
        return ["did not have enviro tab"]*9


# In[93]:

count = 0 # to show how much progress was left
for dirName, subdirlist, fileList in os.walk(questdir):
	total = len(fileList)*1.0
    for fname in fileList:
        if "~$" in fname: #temp files created by excel need to be skipped as it breaks the load workbook method.
            continue
        count += 1
        print "progress: {} working on {}...".format(count/total, fname)
        workbook = opxl.load_workbook(questdir+"\\"+fname)
        #print "Info for building: {}".format(fname)
        #print enviro_data_pull(workbook)
        #print bldg_data_pull(workbook)
        enviro_list.append(enviro_data_pull(workbook))
        bldg_list.append(bldg_data_pull(workbook))
        fname_list.append(fname)



fname_df = pd.DataFrame(fname_list, columns=["filename"])
bldg_df = pd.DataFrame(bldg_list, columns=(
                                "Building Type", # Cell C17
                                'Construction Class', #Cell D17
                                "Exterior Condition" , # Cell E17
                                 ))
enviro_df = pd.DataFrame(enviro_list,columns=("1. During the site assessment were any SUSPECT Asbestos Containing Materials (ACMs) identified? If Yes,                                 provide details below (where?, list what materials)", #Cell H15
                                "Comments for Question 1", #Cell A 16
                                "2. During the site assessment were any issues with Mold identified? If Yes, provide \
                                details below. (where? How much in SF? List all materials)", # Cell H17
                                'Comments for Question 2', #Cell A18
                                "3. During the site assessment were any water infiltration issues identified? \
                                If Yes, provide details below. (where? How much in SF? List all materials)" , # Cell H19
                                'Comments for Question 3' , #Cell A20
                                "4. During the site assessment were any Indoor Air Quality (IAQ) issues identified? \
                                If Yes, provide details below.  (Store manager said HVAC issues? Muggy? What details?)", #Cell H21
                                'Comments for Question 4', #Cell A22
                                "Additional Comments"))

df = pd.concat([fname_df,bldg_df, enviro_df], axis=1) #join vertically

#export to excel :)
writer = pd.ExcelWriter("questionnaire_bldglvl_data"+datenow+".xlsx")
df.to_excel(writer,"Data")
writer.save()
